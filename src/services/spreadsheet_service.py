"""
Serviço de Planilha Unificado
Suporta: Arquivo Local OU OneDrive (quando aprovado)
"""

from typing import List, Dict, Optional
from pathlib import Path
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class SpreadsheetService:
    """Interface unificada para ler/escrever planilhas"""
    
    def __init__(self, mode: str = "local", filepath: str = None):
        """
        Args:
            mode: "local" (arquivo no PC) ou "onedrive" (quando liberado)
            filepath: Caminho para arquivo local (se mode="local")
        """
        self.mode = mode
        self.filepath = filepath or "./uploads/planilha_recebimento.xlsx"
        
        logger.info(f"Serviço iniciado em modo: {mode}")
    
    def read_spreadsheet(self) -> Dict:
        """
        Ler planilha (local ou OneDrive).
        
        Returns:
            {
                'sheet_name': str,
                'headers': list,
                'data': list of dicts,
                'row_count': int,
                'timestamp': datetime
            }
        """
        if self.mode == "local":
            return self._read_local()
        elif self.mode == "onedrive":
            return self._read_onedrive()
        else:
            raise ValueError(f"Mode inválido: {self.mode}")
    
    def _read_local(self) -> Dict:
        """Ler arquivo Excel local"""
        try:
            import openpyxl
            
            logger.info(f"Lendo arquivo local: {self.filepath}")
            
            if not Path(self.filepath).exists():
                logger.error(f"❌ Arquivo não encontrado: {self.filepath}")
                return None
            
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb.active
            
            # Extrair cabeçalho
            headers = [cell.value for cell in ws[1]]
            
            # Extrair dados
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Se tem algum valor na linha
                    data.append(dict(zip(headers, row)))
            
            result = {
                'sheet_name': ws.title,
                'headers': headers,
                'data': data,
                'row_count': len(data),
                'timestamp': datetime.now(),
                'source': 'local'
            }
            
            logger.info(f"✅ Leitura completa: {len(data)} linhas")
            return result
        
        except Exception as e:
            logger.error(f"Erro ao ler arquivo local: {str(e)}")
            return None
    
    def _read_onedrive(self) -> Dict:
        """Ler arquivo do OneDrive (quando aprovado)"""
        logger.warning("OneDrive ainda não aprovado. Use modo='local' por enquanto.")
        return None
    
    def write_data(self, data: List[Dict], sheet_name: str = "Recebimento") -> bool:
        """
        Escrever dados para planilha.
        
        Args:
            data: Lista de dicts com dados
            sheet_name: Nome da aba
        
        Returns:
            True se bem-sucedido
        """
        if self.mode == "local":
            return self._write_local(data, sheet_name)
        elif self.mode == "onedrive":
            return self._write_onedrive(data, sheet_name)
        else:
            raise ValueError(f"Mode inválido: {self.mode}")
    
    def _write_local(self, data: List[Dict], sheet_name: str) -> bool:
        """Escrever dados em arquivo Excel local"""
        try:
            import openpyxl
            
            logger.info(f"Escrevendo para: {self.filepath}")
            
            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not data:
                logger.warning("Sem dados para escrever")
                return False
            
            # Cabeçalho
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Dados
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            # Salvar
            Path(self.filepath).parent.mkdir(parents=True, exist_ok=True)
            wb.save(self.filepath)
            
            logger.info(f"✅ Escrita completa: {len(data)} linhas")
            return True
        
        except Exception as e:
            logger.error(f"Erro ao escrever: {str(e)}")
            return False
    
    def _write_onedrive(self, data: List[Dict], sheet_name: str) -> bool:
        """Escrever para OneDrive (quando aprovado)"""
        logger.warning("OneDrive ainda não aprovado. Use modo='local' por enquanto.")
        return False
    
    def compare_with_remote(self, remote_path: str) -> Dict:
        """
        Comparar planilha local com arquivo remoto.
        
        Args:
            remote_path: Caminho do arquivo de comparação
        
        Returns:
            {
                'added': lista de linhas novas,
                'removed': lista de linhas removidas,
                'modified': lista de linhas modificadas,
                'unchanged': lista de linhas iguais
            }
        """
        local_data = self.read_spreadsheet()
        
        if not local_data:
            logger.error("Erro ao ler dados locais")
            return None
        
        try:
            import openpyxl
            
            logger.info(f"Comparando com: {remote_path}")
            
            wb_remote = openpyxl.load_workbook(remote_path)
            ws_remote = wb_remote.active
            
            remote_headers = [cell.value for cell in ws_remote[1]]
            remote_data = []
            for row in ws_remote.iter_rows(min_row=2, values_only=True):
                if any(row):
                    remote_data.append(dict(zip(remote_headers, row)))
            
            # Comparar
            local_set = set(str(sorted(d.items())) for d in local_data['data'])
            remote_set = set(str(sorted(d.items())) for d in remote_data)
            
            # Resultados
            added = [d for d in local_data['data'] 
                    if str(sorted(d.items())) not in remote_set]
            removed = [d for d in remote_data 
                      if str(sorted(d.items())) not in local_set]
            unchanged = [d for d in local_data['data'] 
                        if str(sorted(d.items())) in remote_set]
            
            result = {
                'added': added,
                'removed': removed,
                'modified': [],  # TODO: Implementar
                'unchanged': unchanged,
                'total_local': len(local_data['data']),
                'total_remote': len(remote_data)
            }
            
            logger.info(f"Comparação: +{len(added)}, -{len(removed)}, ={len(unchanged)}")
            return result
        
        except Exception as e:
            logger.error(f"Erro na comparação: {str(e)}")
            return None
