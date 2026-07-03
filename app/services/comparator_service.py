"""Serviço de Comparação - Compara dados extraídos entre duas versões de documento"""

from openpyxl.utils import get_column_letter
from app.models import db, Comparison, DocumentVersion


class ComparatorService:
    
    @staticmethod
    def compare_versions(new_version_id, old_version_id=None):
        """
        Compara a nova versão com uma versão anterior (geralmente a última aprovada)
        
        Args:
            new_version_id: ID da nova versão (DocumentVersion)
            old_version_id: ID da versão anterior (DocumentVersion), ou None se for a primeira versão
            
        Returns:
            {
                'differences_count': int,
                'differences': [
                    {
                        'sheet': 'Planilha1',
                        'row': 5,
                        'col': 'A',
                        'old_value': 'valor_antigo',
                        'new_value': 'valor_novo'
                    }
                ],
                'comparison_id': str
            }
        """
        try:
            new_version = DocumentVersion.query.get(new_version_id)
            if not new_version:
                raise ValueError("Nova versão não encontrada")
                
            differences = []
            if not old_version_id:
                comparison = Comparison(
                    document_version_id=new_version_id,
                    compare_version_id=None,
                    differences_count=0,
                    differences_data={'differences': []}
                )
                db.session.add(comparison)
                db.session.commit()
                
                return {
                    'differences_count': 0,
                    'differences': [],
                    'comparison_id': comparison.id
                }
                
            old_version = DocumentVersion.query.get(old_version_id)
            if not old_version:
                raise ValueError("Versão anterior não encontrada")
                
            extracted_data = new_version.extracted_data or {}
            old_data = old_version.extracted_data or {}
            new_sheets = set(extracted_data.keys())
            old_sheets = set(old_data.keys())
            all_sheets = new_sheets | old_sheets
            
            for sheet_name in all_sheets:
                if sheet_name not in old_sheets:
                    new_rows = extracted_data[sheet_name]
                    for r_idx, row in enumerate(new_rows):
                        for c_idx, cell in enumerate(row):
                            if cell is not None and cell != '':
                                differences.append({
                                    'sheet': sheet_name,
                                    'row': r_idx + 1,
                                    'col': get_column_letter(c_idx + 1),
                                    'old_value': '',
                                    'new_value': str(cell)
                                })
                    continue
                if sheet_name not in new_sheets:
                    old_rows = old_data[sheet_name]
                    for r_idx, row in enumerate(old_rows):
                        for c_idx, cell in enumerate(row):
                            if cell is not None and cell != '':
                                differences.append({
                                    'sheet': sheet_name,
                                    'row': r_idx + 1,
                                    'col': get_column_letter(c_idx + 1),
                                    'old_value': str(cell),
                                    'new_value': ''
                                })
                    continue
                new_rows = extracted_data[sheet_name]
                old_rows = old_data[sheet_name]
                
                max_rows = max(len(new_rows), len(old_rows))
                
                for row_idx in range(max_rows):
                    new_row = new_rows[row_idx] if row_idx < len(new_rows) else []
                    old_row = old_rows[row_idx] if row_idx < len(old_rows) else []
                    
                    max_cols = max(len(new_row), len(old_row))
                    
                    for col_idx in range(max_cols):
                        new_val = new_row[col_idx] if col_idx < len(new_row) else None
                        old_val = old_row[col_idx] if col_idx < len(old_row) else None
                        
                        if new_val != old_val:
                            if (new_val is None or new_val == '') and (old_val is None or old_val == ''):
                                continue
                                
                            differences.append({
                                'sheet': sheet_name,
                                'row': row_idx + 1,
                                'col': get_column_letter(col_idx + 1),
                                'old_value': str(old_val) if old_val is not None else '',
                                'new_value': str(new_val) if new_val is not None else ''
                            })
            
            comparison = Comparison(
                document_version_id=new_version_id,
                compare_version_id=old_version_id,
                differences_count=len(differences),
                differences_data={
                    'differences': differences,
                    'new_sheets': list(new_sheets),
                    'old_sheets': list(old_sheets)
                }
            )
            db.session.add(comparison)
            db.session.commit()
            
            return {
                'differences_count': len(differences),
                'differences': differences,
                'comparison_id': comparison.id
            }
            
        except Exception as e:
            db.session.rollback()
            raise Exception(f"Erro ao comparar dados: {str(e)}")
            
    @staticmethod
    def get_comparison(comparison_id):
        """Retorna detalhes de uma comparação"""
        return Comparison.query.get(comparison_id)
    
    @staticmethod
    def format_differences_for_display(differences):
        """Formata diferenças para exibição em HTML"""
        by_sheet = {}
        for diff in differences:
            sheet = diff.get('sheet', 'Geral')
            if sheet not in by_sheet:
                by_sheet[sheet] = []
            by_sheet[sheet].append(diff)
        return by_sheet
