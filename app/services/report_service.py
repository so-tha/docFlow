"""Serviço de Relatórios - Upload, processamento e comparação sem salvar arquivo"""

import hashlib
from werkzeug.utils import secure_filename
from flask import current_app
from app.models import db, Report, Comparison, User
from datetime import datetime
from dateutil.parser import parse as parse_date


class ReportService:
    
    ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
    
    @staticmethod
    def allowed_file(filename):
        """Valida extensão do arquivo"""
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ReportService.ALLOWED_EXTENSIONS
    
    @staticmethod
    def _make_json_serializable(obj):
        """
        Converte objetos para tipos JSON-serializáveis
        Handle: datetime, date, time, etc.
        """
        from datetime import datetime, date, time
        
        if isinstance(obj, (datetime, date, time)):
            return obj.isoformat()
        elif isinstance(obj, (list, tuple)):
            return [ReportService._make_json_serializable(item) for item in obj]
        elif isinstance(obj, dict):
            return {key: ReportService._make_json_serializable(value) for key, value in obj.items()}
        else:
            return obj
    
    @staticmethod
    def extract_file_data(file):
        """
        Extrai dados do arquivo sem salvar localmente
        
        Suporta: XLSX, XLS, CSV
        Retorna: {sheet_name: [rows_data]} com todos os tipos JSON-serializáveis
        """
        try:
            import openpyxl
            import csv
            from io import StringIO
            
            file_ext = file.filename.rsplit('.', 1)[1].lower()
            file_data = file.read()
            
            if file_ext in ['xlsx', 'xls']:
                # Ler Excel
                from io import BytesIO
                workbook = openpyxl.load_workbook(BytesIO(file_data))
                extracted = {}
                
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    sheet_data = []
                    
                    for row in worksheet.iter_rows(values_only=True):
                        # ✅ Converter para JSON-serializable
                        clean_row = [ReportService._make_json_serializable(cell) for cell in row]
                        sheet_data.append(clean_row)
                    
                    extracted[sheet_name] = sheet_data
                
                return extracted
            
            elif file_ext == 'csv':
                # Ler CSV
                text_data = file_data.decode('utf-8')
                reader = csv.reader(StringIO(text_data))
                rows = [row for row in reader]
                return {'data': rows}
        
        except Exception as e:
            raise ValueError(f"Erro ao extrair dados do arquivo: {str(e)}")
    
    @staticmethod
    def process_uploaded_file(file, user_id, sharepoint_service=None):
        """
        Processa arquivo enviado: extrai dados, compara com SharePoint, armazena
        
        NÃO salva arquivo localmente - apenas armazena dados em JSON no banco
        
        Deduplicação: se hash já existe, retorna o relatório existente
        
        Args:
            file: Arquivo enviado (werkzeug FileStorage)
            user_id: ID do usuário que fez upload
            sharepoint_service: Serviço para comparação com SharePoint
        
        Returns:
            (Report, comparison_result, is_duplicate)
            - Report object com dados extraídos
            - Comparação realizada
            - Flag indicando se era duplicata
        """
        try:
            if not ReportService.allowed_file(file.filename):
                raise ValueError(f"Tipo de arquivo não permitido: {file.filename}")
            
            # Preparar dados
            original_filename = secure_filename(file.filename)
            file.seek(0)
            file_data = file.read()
            file_hash = hashlib.sha256(file_data).hexdigest()
            file_size = len(file_data)
            
            # ✅ VERIFICAR DUPLICATA pelo hash
            existing_report = Report.query.filter_by(file_hash=file_hash).first()
            
            if existing_report:
                # Arquivo já foi enviado antes
                # Registra novo upload na auditoria, mas não duplica no banco
                return existing_report, {
                    'differences_count': 0,
                    'differences_data': {'message': 'Arquivo já foi processado'},
                    'comparison_id': existing_report.comparisons.first().id if existing_report.comparisons.first() else None
                }, True
            
            # Extrair dados do arquivo
            file.seek(0)
            extracted_data = ReportService.extract_file_data(file)
            
            # ✅ CRIAR RECORD NO BANCO - TUDO EM JSON
            report = Report(
                original_filename=original_filename,
                file_size=file_size,
                file_hash=file_hash,
                extracted_data=extracted_data,  # JSON com dados completos
                uploaded_by_id=user_id,
                status='pending'
            )
            
            db.session.add(report)
            db.session.flush()  # Obter ID antes de fazer comparação
            
            # COMPARAR com SharePoint (se serviço disponível)
            if sharepoint_service:
                comparison_result = ReportService.compare_with_sharepoint(
                    report.id,
                    extracted_data,
                    sharepoint_service
                )
            else:
                comparison_result = {
                    'differences_count': 0,
                    'differences_data': {
                        'sheets': list(extracted_data.keys()) if extracted_data else [],
                        'status': 'pending'
                    },
                    'comparison_id': None
                }
            
            db.session.commit()
            return report, comparison_result, False  # False = não era duplicata
        
        except Exception as e:
            db.session.rollback()
            raise Exception(f"Erro ao processar arquivo: {str(e)}")
    
    @staticmethod
    def compare_with_sharepoint(report_id, extracted_data, sharepoint_service):
        """
        Compara dados extraídos com SharePoint e armazena comparação
        
        Returns:
            {differences_count, differences_data, comparison_id}
        """
        try:
            # TODO: Integrar com SharePointService
            # Por enquanto, mock da comparação
            
            comparison = Comparison(
                report_id=report_id,
                differences_count=0,
                differences_data={
                    'status': 'pending',
                    'sheets': list(extracted_data.keys())
                }
            )
            
            db.session.add(comparison)
            db.session.commit()
            
            return {
                'differences_count': comparison.differences_count,
                'differences_data': comparison.differences_data,
                'comparison_id': comparison.id
            }
        
        except Exception as e:
            raise Exception(f"Erro ao comparar com SharePoint: {str(e)}")
    
    @staticmethod
    def get_pending_reports(limit=50):
        """Retorna relatórios pendentes de aprovação"""
        return Report.query.filter_by(status='pending').order_by(
            Report.created_at.desc()
        ).limit(limit).all()
    
    @staticmethod
    def get_all_reports(limit=50, offset=0):
        """Retorna todos os relatórios com paginação"""
        return Report.query.order_by(
            Report.created_at.desc()
        ).limit(limit).offset(offset).all()
    
    @staticmethod
    def get_report_by_user(user_id, limit=50):
        """Retorna relatórios enviados por um usuário específico"""
        return Report.query.filter_by(uploaded_by_id=user_id).order_by(
            Report.created_at.desc()
        ).limit(limit).all()
    
    @staticmethod
    def extract_month_from_data(extracted_data):
        """
        Extrai o mês dos dados pela coluna 'DATA EMISSÃO'
        
        Args:
            extracted_data: {sheet_name: [rows]} extraído do arquivo
        
        Returns:
            {'month': int, 'year': int, 'sheet_name': str}
            Ex: {'month': 4, 'year': 2026, 'sheet_name': 'Planilha1'}
            
            Return None se não conseguir detectar
        """
        try:
            # Procurar coluna "DATA EMISSÃO" na primeira sheet
            for sheet_name, rows in extracted_data.items():
                if not rows or len(rows) < 2:
                    continue
                
                # Header está na primeira linha
                header = rows[0]
                
                # Procurar coluna "DATA EMISSÃO"
                data_col_idx = None
                for idx, col_name in enumerate(header):
                    if col_name and 'DATA' in str(col_name).upper() and 'EMISSÃO' in str(col_name).upper():
                        data_col_idx = idx
                        break
                
                if data_col_idx is None:
                    continue
                
                # Procurar primeira data válida nos dados (pulando header)
                for row in rows[1:]:
                    if len(row) > data_col_idx:
                        date_value = row[data_col_idx]
                        
                        if date_value:
                            try:
                                # Tentar parse da data
                                if isinstance(date_value, str):
                                    parsed_date = parse_date(date_value, dayfirst=True)
                                else:
                                    parsed_date = date_value
                                
                                return {
                                    'month': parsed_date.month,
                                    'year': parsed_date.year,
                                    'sheet_name': sheet_name,
                                    'date_obj': parsed_date
                                }
                            except:
                                continue
            
            return None
        
        except Exception as e:
            raise ValueError(f"Erro ao extrair mês dos dados: {str(e)}")
    
    @staticmethod
    def get_sheet_name_for_month(month, year):
        """
        Mapeia número do mês para nome da aba do SharePoint
        
        Ex: (4, 2026) → "ABR 26"
        
        Args:
            month: Número do mês (1-12)
            year: Ano (ex: 2026)
        
        Returns:
            Nome da aba no SharePoint (ex: "ABR 26")
        """
        months_map = {
            1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR",
            5: "MAI", 6: "JUN", 7: "JUL", 8: "AGO",
            9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ"
        }
        
        month_name = months_map.get(month, "UNKNOWN")
        year_short = str(year)[-2:]  # Pega últimos 2 dígitos
        
        return f"{month_name} {year_short}"
    
    @staticmethod
    def get_reports_by_status(status, limit=50):
        """Retorna relatórios por status específico"""
        return Report.query.filter_by(status=status).order_by(
            Report.created_at.desc()
        ).limit(limit).all()
    
    @staticmethod
    def approve_report(report_id, user_id, comment=''):
        """
        Aprova relatório e o BLOQUEIA para posteriores modificações
        Se aprovado, sube as novas linhas para SharePoint
        """
        try:
            report = Report.query.get(report_id)
            if not report:
                raise ValueError("Relatório não encontrado")
            
            if report.is_locked:
                raise ValueError("Relatório já foi decidido e não pode ser modificado")
            
            report.status = 'approved'
            report.approved_at = datetime.utcnow()
            report.decision_by_id = user_id
            report.decision_comment = comment
            report.lock()  # BLOQUEIA para posteriores mudanças
            
            db.session.commit()
            
            # ✅ SINCRONIZAR com SharePoint
            from app.services.sharepoint_service import SharePointService
            try:
                sharepoint_service = SharePointService()
                sharepoint_service.sync_approved_report(report_id, report.extracted_data)
            except Exception as e:
                # Log erro mas não falha a aprovação
                print(f"Aviso: Erro ao sincronizar com SharePoint: {str(e)}")
            
            return report
        except Exception as e:
            db.session.rollback()
            raise Exception(f"Erro ao aprovar relatório: {str(e)}")
    
    @staticmethod
    def reject_report(report_id, user_id, comment=''):
        """
        Rejeita relatório e o BLOQUEIA para posteriores modificações
        Uma vez rejeitado, não pode ser reaprovado
        """
        try:
            report = Report.query.get(report_id)
            if not report:
                raise ValueError("Relatório não encontrado")
            
            if report.is_locked:
                raise ValueError("Relatório já foi decidido e não pode ser modificado")
            
            report.status = 'rejected'
            report.rejected_at = datetime.utcnow()
            report.decision_by_id = user_id
            report.decision_comment = comment
            report.lock()  # BLOQUEIA para posteriores mudanças
            
            db.session.commit()
            return report
        except Exception as e:
            db.session.rollback()
            raise Exception(f"Erro ao rejeitar relatório: {str(e)}")
    
    @staticmethod
    def format_report(report):
        """Formata relatório para JSON"""
        uploader = User.query.get(report.uploaded_by_id)
        approver = User.query.get(report.decision_by_id) if report.decision_by_id else None
        
        return {
            'id': report.id,
            'original_filename': report.original_filename,
            'file_size': report.file_size,
            'status': report.status,
            'is_locked': report.is_locked,  # Relatório bloqueado após decisão
            'uploaded_by': uploader.name if uploader else 'Desconhecido',
            'uploaded_by_email': uploader.email if uploader else 'N/A',
            'uploaded_at': report.created_at.isoformat(),
            'approved_by': approver.name if approver else None,
            'approved_at': report.approved_at.isoformat() if report.approved_at else None,
            'rejected_at': report.rejected_at.isoformat() if report.rejected_at else None,
            'decision_comment': report.decision_comment,
            'comparison_id': report.comparisons.first().id if report.comparisons.first() else None,
            'extracted_data_summary': {
                'sheets': list(report.extracted_data.keys()) if report.extracted_data else []
            }
        }
