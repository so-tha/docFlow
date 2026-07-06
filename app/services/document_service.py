import hashlib
from werkzeug.utils import secure_filename
from flask import current_app
from app.models import db, User, Document, DocumentVersion, Comparison
from app.services.comparator_service import ComparatorService
from datetime import datetime
import openpyxl
import csv
from io import StringIO, BytesIO

class DocumentService:
    ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

    @staticmethod
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in DocumentService.ALLOWED_EXTENSIONS

    @staticmethod
    def _make_json_serializable(obj):
        from datetime import datetime, date, time
        if isinstance(obj, (datetime, date, time)):
            return obj.isoformat()
        elif isinstance(obj, (list, tuple)):
            return [DocumentService._make_json_serializable(item) for item in obj]
        elif isinstance(obj, dict):
            return {key: DocumentService._make_json_serializable(value) for key, value in obj.items()}
        else:
            return obj

    @staticmethod
    def extract_file_data(file):
        try:
            file_ext = file.filename.rsplit('.', 1)[1].lower()
            file_data = file.read()
            if file_ext in ['xlsx', 'xls']:
                workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
                extracted = {}
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    sheet_data = []
                    for row in worksheet.iter_rows(values_only=True):
                        clean_row = [DocumentService._make_json_serializable(cell) for cell in row]
                        sheet_data.append(clean_row)
                    extracted[sheet_name] = sheet_data
                return extracted
            elif file_ext == 'csv':
                try:
                    text_data = file_data.decode('utf-8')
                except UnicodeDecodeError:
                    text_data = file_data.decode('latin1')
                reader = csv.reader(StringIO(text_data))
                rows = [row for row in reader]
                return {'data': rows}
        except Exception as e:
            raise ValueError(f'Erro ao extrair dados do arquivo: {str(e)}')

    @staticmethod
    def process_uploaded_file(file, user_id, document_id=None, document_name=None):
        try:
            if not DocumentService.allowed_file(file.filename):
                raise ValueError(f'Tipo de arquivo não permitido: {file.filename}')
            original_filename = secure_filename(file.filename)
            file.seek(0)
            file_data = file.read()
            file_hash = hashlib.sha256(file_data).hexdigest()
            file_size = len(file_data)
            existing_version = DocumentVersion.query.filter_by(file_hash=file_hash).first()
            if existing_version:
                comparison = existing_version.comparisons.first()
                return (existing_version, {'differences_count': comparison.differences_count if comparison else 0, 'comparison_id': comparison.id if comparison else None}, True)
            file.seek(0)
            extracted_data = DocumentService.extract_file_data(file)
            if document_id:
                document = Document.query.get(document_id)
                if not document:
                    raise ValueError('Documento selecionado não encontrado')
            else:
                if not document_name:
                    document_name = original_filename.rsplit('.', 1)[0].replace('_', ' ').title()
                document = Document(name=document_name, created_by_id=user_id)
                db.session.add(document)
                db.session.flush()
            latest_version = document.get_latest_version()
            version_number = 1 if not latest_version else latest_version.version_number + 1
            status = 'approved' if version_number == 1 else 'pending'
            is_locked = True if version_number == 1 else False
            new_version = DocumentVersion(document_id=document.id, version_number=version_number, original_filename=original_filename, file_size=file_size, file_hash=file_hash, extracted_data=extracted_data, uploaded_by_id=user_id, status=status, is_locked=is_locked)
            if version_number == 1:
                new_version.approved_at = datetime.utcnow()
                new_version.decision_by_id = user_id
                new_version.decision_comment = 'Versão inicial (Base)'
            db.session.add(new_version)
            db.session.flush()
            comparison_result = None
            if version_number > 1:
                latest_approved = document.get_latest_approved_version()
                old_version_id = latest_approved.id if latest_approved else None
                comparison_result = ComparatorService.compare_versions(new_version.id, old_version_id)
            else:
                comparison_result = ComparatorService.compare_versions(new_version.id, None)
            db.session.commit()
            return (new_version, comparison_result, False)
        except Exception as e:
            db.session.rollback()
            raise Exception(f'Erro ao processar arquivo: {str(e)}')

    @staticmethod
    def approve_version(version_id, user_id, comment=''):
        try:
            version = DocumentVersion.query.get(version_id)
            if not version:
                raise ValueError('Versão de documento não encontrada')
            if version.is_locked:
                raise ValueError('Esta versão já foi decidida')
            version.status = 'approved'
            version.approved_at = datetime.utcnow()
            version.decision_by_id = user_id
            version.decision_comment = comment
            version.lock()
            version.document.updated_at = datetime.utcnow()
            db.session.commit()
            return version
        except Exception as e:
            db.session.rollback()
            raise Exception(f'Erro ao aprovar versão: {str(e)}')

    @staticmethod
    def reject_version(version_id, user_id, comment=''):
        try:
            version = DocumentVersion.query.get(version_id)
            if not version:
                raise ValueError('Versão de documento não encontrada')
            if version.is_locked:
                raise ValueError('Esta versão já foi decidida')
            version.status = 'rejected'
            version.rejected_at = datetime.utcnow()
            version.decision_by_id = user_id
            version.decision_comment = comment
            version.lock()
            db.session.commit()
            return version
        except Exception as e:
            db.session.rollback()
            raise Exception(f'Erro ao rejeitar versão: {str(e)}')

    @staticmethod
    def get_documents_by_user(user_id):
        return Document.query.filter_by(created_by_id=user_id).order_by(Document.updated_at.desc()).all()

    @staticmethod
    def get_all_documents():
        return Document.query.order_by(Document.updated_at.desc()).all()

    @staticmethod
    def format_version(version):
        uploader = User.query.get(version.uploaded_by_id)
        approver = User.query.get(version.decision_by_id) if version.decision_by_id else None
        comparison = version.comparisons.first()
        return {'id': version.id, 'document_id': version.document_id, 'document_name': version.document.name, 'version_number': version.version_number, 'original_filename': version.original_filename, 'file_size': version.file_size, 'status': version.status, 'is_locked': version.is_locked, 'uploaded_by': uploader.name if uploader else 'Desconhecido', 'uploaded_by_email': uploader.email if uploader else 'N/A', 'uploaded_at': version.created_at.strftime('%d/%m/%Y %H:%M:%S'), 'approved_by': approver.name if approver else None, 'approved_at': version.approved_at.strftime('%d/%m/%Y %H:%M:%S') if version.approved_at else None, 'rejected_at': version.rejected_at.strftime('%d/%m/%Y %H:%M:%S') if version.rejected_at else None, 'decision_comment': version.decision_comment, 'differences_count': comparison.differences_count if comparison else 0, 'comparison_id': comparison.id if comparison else None}